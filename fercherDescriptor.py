import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configuarion
PAT = "ghp_yourpersonalaccesstokenhere"
organizations = ["Org1", "Org2"]

session = requests.Session()
session.auth = ("", PAT)

# Fetch projects for an organization
def getProjects(org):
    url = f"https://dev.azure.com/{org}/_apis/projects?api-version=7.1"
    response = session.get(url)
    response.raise_for_status()
    return response.json()["value"]

# Fetch descriptor of a project
def getprojectDescriptors(org, projectId):
    url = f"https://vssps.dev.azure.com/{org}/_apis/graph/descriptors/{projectId}?api-version=7.1-preview.1" # preview is optional, check whats working for you
    response = session.get(url)
    response.raise_for_status()
    return response.json()["value"]

# Fetch groups under a given project descriptor
def getGroups(org, pdescriptor):
    url = f"https://vssps.dev.azure.com/{org}/_apis/graph/groups?scopeDescriptor={pdescriptor}&api-version=7.1-preview.1"
    response = session.get(url)
    response.raise_for_status()
    return response.json().get("value", [])

# Autofit column widths in Excel worksheet
def autofit(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                l = len(str(cell.value))
                if l > max_length:
                    max_length = l
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

# Main logic
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "All Project Groups"
    ws.append(["Organization", "Project Name", "Group Name", "Group Descriptor"])

    for org in organizations:
        projects = getProjects(org)
        for project in projects:
            projectId = project["id"]
            projectName = project["name"]
            pdescriptor = getprojectDescriptors(org, projectId)
            groups = getGroups(org, pdescriptor)

            for group in groups:
                groupName = group["displayName"]
                groupDescriptor = group["descriptor"]
                ws.append([org, projectName, groupName, groupDescriptor]) # or you can use group.get("propertyName", "") for additional properties

    autofit(ws)
    wb.save("ProjectGroups.xlsx")
    print("Data saved to ProjectGroups.xlsx")

if __name__ == "__main__":
    main()